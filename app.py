from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file, jsonify
from dotenv import load_dotenv
from datetime import date, datetime
import os
import database
from pytz import timezone
from openpyxl import Workbook
from io import BytesIO
from werkzeug.security import generate_password_hash, check_password_hash
import json
import requests
from urllib.parse import urlencode
from openai import OpenAI

load_dotenv()

app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY", "fallback-secret")

database.init_db()

# Google OAuth Configuration
GOOGLE_CLIENT_ID = os.getenv("GOOGLE_CLIENT_ID", "")
GOOGLE_CLIENT_SECRET = os.getenv("GOOGLE_CLIENT_SECRET", "")
GOOGLE_REDIRECT_URI = os.getenv("GOOGLE_REDIRECT_URI", "http://localhost:5000/auth/google/callback")
GOOGLE_AUTH_URI = "https://accounts.google.com/o/oauth2/v2/auth"
GOOGLE_TOKEN_URI = "https://oauth2.googleapis.com/token"
GOOGLE_USERINFO_URI = "https://openidconnect.googleapis.com/v1/userinfo"


def login_required(f):
    from functools import wraps

    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated


@app.route("/", methods=["GET"])
@login_required
def index():
    user_id = session["user_id"]
    user = database.get_user_by_id(user_id)
    india = timezone("Asia/Kolkata")
    selected_date = request.args.get(
        "date", datetime.now(india).date().isoformat())
    entries = database.get_entries_by_date(user_id, selected_date)
    # Convert Row objects to dictionaries for JSON serialization
    entries = [dict(entry) for entry in entries]
    total = database.get_total_calories(user_id, selected_date)
    all_dates = database.get_all_dates(user_id)
    today = datetime.now(india).date().isoformat()

    # Ensure today is always in the date list for the sidebar
    if today not in all_dates:
        all_dates = [today] + list(all_dates)

    return render_template(
        "dashboard.html",
        entries=entries,
        selected_date=selected_date,
        total=total,
        all_dates=all_dates,
        today=today,
        username=user["username"] if user else "User",
        is_oauth_user=session.get("auth_method") == "google"
    )


@app.route("/add", methods=["POST"])
@login_required
def add():
    user_id = session["user_id"]
    food_item = request.form.get("food_item", "").strip()
    calories = request.form.get("calories", "").strip()
    entry_date = request.form.get("entry_date", date.today().isoformat())

    if not food_item or not calories:
        flash("Both food item and calories are required.", "error")
        return redirect(url_for("index", date=entry_date))

    try:
        calories = int(calories)
        if calories <= 0:
            raise ValueError
    except ValueError:
        flash("Calories must be a positive number.", "error")
        return redirect(url_for("index", date=entry_date))

    database.add_entry(user_id, entry_date, food_item, calories)
    return redirect(url_for("index", date=entry_date))


@app.route("/delete/<int:entry_id>", methods=["POST"])
@login_required
def delete(entry_id):
    user_id = session["user_id"]
    entry_date = request.form.get("entry_date", date.today().isoformat())
    database.delete_entry(entry_id, user_id)
    return redirect(url_for("index", date=entry_date))


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()
        
        user = database.get_user_by_username(username)
        
        if user and check_password_hash(user["password_hash"], password):
            session["user_id"] = user["id"]
            flash(f"Welcome back, {user['username']}!", "success")
            return redirect(url_for("index"))
        
        flash("Invalid username or password.", "error")
    
    return render_template("login.html")


@app.route("/signup")
def signup():
    """Redirect to Google OAuth signup"""
    flash("Please sign up using Google to get started.", "info")
    return redirect(url_for("login"))


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/export", methods=["GET"])
@login_required
def export():
    user_id = session["user_id"]
    # Get all dates from database for this user
    all_dates = database.get_all_dates(user_id)

    if not all_dates:
        flash("No data to export.", "error")
        return redirect(url_for("index"))

    # Create a workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    from openpyxl.styles import Font, PatternFill

    # Create a sheet for each date
    for export_date in all_dates:
        entries = database.get_entries_by_date(user_id, export_date)
        total = database.get_total_calories(user_id, export_date)

        # Create sheet with date as name
        ws = wb.create_sheet(title=export_date)

        # Add headers
        headers = ["Food Item", "Calories", "Time"]
        ws.append(headers)

        # Style headers
        header_fill = PatternFill(start_color="FF8B5000",
                                  end_color="FF8B5000", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        # Add data rows
        for entry in entries:
            ws.append([
                entry["food_item"],
                entry["calories"],
                entry["created_at"][:16]
            ])

        # Add total row
        ws.append(["Total", total, ""])
        total_fill = PatternFill(start_color="FFFF9E64",
                                 end_color="FFFF9E64", fill_type="solid")
        total_font = Font(bold=True)
        for cell in ws[ws.max_row]:
            cell.fill = total_fill
            cell.font = total_font

        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 18

    # Save to BytesIO and send
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    filename = "calories_all_dates.xlsx"
    return send_file(excel_file, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name=filename)


@app.route("/change-password", methods=["POST"])
@login_required
def change_password():
    """Handle password change request"""
    user_id = session["user_id"]
    user = database.get_user_by_id(user_id)
    is_oauth_user = session.get("auth_method") == "google"
    
    current_password = request.form.get("current_password", "").strip()
    new_password = request.form.get("new_password", "").strip()
    confirm_password = request.form.get("confirm_password", "").strip()
    
    # Validation
    if not new_password or not confirm_password:
        flash("Password fields are required.", "error")
        return redirect(url_for("index"))
    
    # For non-OAuth users, verify current password
    if not is_oauth_user:
        if not current_password:
            flash("Current password is required.", "error")
            return redirect(url_for("index"))
        
        # Verify current password
        if not check_password_hash(user["password_hash"], current_password):
            flash("Current password is incorrect.", "error")
            return redirect(url_for("index"))
    
    # Check password length
    if len(new_password) < 6:
        flash("New password must be at least 6 characters.", "error")
        return redirect(url_for("index"))
    
    # Check passwords match
    if new_password != confirm_password:
        flash("New passwords do not match.", "error")
        return redirect(url_for("index"))
    
    # Check if new password is same as old password (only for non-OAuth users)
    if not is_oauth_user and check_password_hash(user["password_hash"], new_password):
        flash("New password must be different from current password.", "error")
        return redirect(url_for("index"))
    
    # Update password in database
    password_hash = generate_password_hash(new_password)
    success = database.update_user_password(user_id, password_hash)
    
    if success:
        flash("Password changed successfully!", "success")
    else:
        flash("Failed to change password. Please try again.", "error")
    
    return redirect(url_for("index"))


@app.route("/auth/google")
def auth_google():
    """Initiate Google OAuth flow"""
    if not GOOGLE_CLIENT_ID:
        flash("Google authentication is not configured.", "error")
        return redirect(url_for("login"))
    
    # Generate authorization URL
    auth_params = {
        "client_id": GOOGLE_CLIENT_ID,
        "redirect_uri": GOOGLE_REDIRECT_URI,
        "response_type": "code",
        "scope": "openid email profile",
        "access_type": "offline",
        "prompt": "consent"
    }
    
    auth_url = f"{GOOGLE_AUTH_URI}?{urlencode(auth_params)}"
    return redirect(auth_url)


@app.route("/auth/google/callback")
def auth_google_callback():
    """Handle Google OAuth callback"""
    code = request.args.get("code")
    error = request.args.get("error")
    
    if error:
        flash("Google authentication failed.", "error")
        return redirect(url_for("login"))
    
    if not code:
        flash("No authorization code received.", "error")
        return redirect(url_for("login"))
    
    if not GOOGLE_CLIENT_ID or not GOOGLE_CLIENT_SECRET:
        flash("Google authentication is not configured.", "error")
        return redirect(url_for("login"))
    
    try:
        # Exchange code for token
        token_data = {
            "code": code,
            "client_id": GOOGLE_CLIENT_ID,
            "client_secret": GOOGLE_CLIENT_SECRET,
            "redirect_uri": GOOGLE_REDIRECT_URI,
            "grant_type": "authorization_code"
        }
        
        token_response = requests.post(GOOGLE_TOKEN_URI, data=token_data, timeout=10)
        
        if token_response.status_code != 200:
            flash("Failed to get token from Google.", "error")
            return redirect(url_for("login"))
        
        tokens = token_response.json()
        access_token = tokens.get("access_token")
        
        if not access_token:
            flash("Failed to get access token.", "error")
            return redirect(url_for("login"))
        
        # Get user info
        user_info_response = requests.get(
            GOOGLE_USERINFO_URI,
            headers={"Authorization": f"Bearer {access_token}"},
            timeout=10
        )
        
        if user_info_response.status_code != 200:
            flash("Failed to get user information from Google.", "error")
            return redirect(url_for("login"))
        
        user_info = user_info_response.json()
        email = user_info.get("email")
        name = user_info.get("name", "User")
        
        if not email:
            flash("Could not get email from Google.", "error")
            return redirect(url_for("login"))
        
        # Find or create user
        user = database.get_user_by_email(email)
        
        if user:
            # Existing user, log them in
            session["user_id"] = user["id"]
            session["auth_method"] = "google"
            flash(f"Welcome back, {user['username']}!", "success")
            return redirect(url_for("index"))
        else:
            # New user, create account
            # Generate username from email (first part) with unique suffix if needed
            base_username = email.split("@")[0]
            username = base_username
            counter = 1
            
            # Ensure username is different from email and is unique
            while database.get_user_by_username(username) or username == email:
                username = f"{base_username}{counter}"
                counter += 1
            
            # Create user with Google OAuth (no password)
            password_hash = generate_password_hash(os.urandom(24).hex())
            user_id = database.create_user(username, email, password_hash)
            
            if user_id:
                session["user_id"] = user_id
                session["auth_method"] = "google"
                flash(f"Welcome, {username}! Your account has been created via Google Sign-In.", "success")
                return redirect(url_for("index"))
            else:
                flash("Failed to create user account.", "error")
                return redirect(url_for("login"))
    
    except requests.exceptions.RequestException as e:
        flash("Network error during authentication.", "error")
        return redirect(url_for("login"))
    except Exception as e:
        flash("An error occurred during authentication.", "error")
        return redirect(url_for("login"))


@app.route("/suggest-food", methods=["POST"])
@login_required
def suggest_food():
    """AI-powered food suggestion with minimal token consumption."""
    try:
        food_input = request.json.get("food_input", "").strip()
        
        if not food_input:
            return jsonify({"error": "Food input is required"}), 400
        
        # Initialize OpenAI client
        api_key = os.getenv("OPENAI_API_KEY")
        if not api_key:
            return jsonify({"error": "OpenAI API key not configured"}), 500
        
        client = OpenAI(api_key=api_key)
        
        # Minimal prompt to reduce token consumption
        system_prompt = "You are a calorie estimation assistant. Respond with ONLY: 'food_name: calories_number'. Be brief and accurate."
        
        # Call OpenAI API with gpt-3.5-turbo for cost efficiency
        response = client.chat.completions.create(
            model="gpt-3.5-turbo",
            messages=[
                {
                    "role": "system",
                    "content": system_prompt
                },
                {
                    "role": "user",
                    "content": f"Estimate calories for: {food_input}"
                }
            ],
            temperature=0,  # Deterministic for faster response
            max_tokens=20   # Very short response only
        )
        

        # Extract and parse the response
        suggestion = response.choices[0].message.content.strip()
        print(f"AI Suggestion: {suggestion}")  # Debug log
        
        # Parse response format: "food_name: calories_number"
        if ":" in suggestion:
            parts = suggestion.split(":")
            food_name = parts[0].strip()
            try:
                calories = int(parts[-1].strip().split()[0])  # Extract number even if there's extra text
                return jsonify({
                    "food": food_name,
                    "calories": calories,
                    "suggestion": suggestion
                })
            except (ValueError, IndexError):
                return jsonify({
                    "error": "Could not parse calorie value from response",
                    "raw_response": suggestion
                }), 400
        else:
            return jsonify({
                "error": "Invalid response format",
                "raw_response": suggestion
            }), 400
    
    except Exception as e:
        print(f"Error in suggest_food: {str(e)}")
        return jsonify({"error": str(e)}), 500


if __name__ == "__main__":
    app.run(debug=True)
