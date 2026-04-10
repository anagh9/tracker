"""
Calorie Tracker Routes Blueprint
"""

from flask import send_file
from datetime import datetime
import re

from flask import Blueprint, render_template, request, redirect, url_for, session, flash, send_file, jsonify, Response
from datetime import date
from pytz import timezone
from openpyxl import Workbook
from io import BytesIO
from openai import OpenAI
import os
import database
from utils import login_required, get_timezone, get_today_iso
from config import Config

calorie_bp = Blueprint('calorie', __name__)


@calorie_bp.route("/", methods=["GET"])
@login_required
def index():
    """Calorie tracker dashboard"""
    user_id = session["user_id"]
    user = database.get_user_by_id(user_id)
    tz = get_timezone()
    selected_date = request.args.get("date", date.today().isoformat())
    entries = database.get_entries_by_date(user_id, selected_date)
    entries = [dict(entry) for entry in entries]
    total = database.get_total_calories(user_id, selected_date)
    all_dates = database.get_all_dates(user_id)
    today = date.today().isoformat()

    if today not in all_dates:
        all_dates = [today] + list(all_dates)

    return render_template(
        "calorie/dashboard.html",
        entries=entries,
        selected_date=selected_date,
        total=total,
        all_dates=all_dates,
        today=today,
        username=user["username"] if user else "User",
        is_oauth_user=session.get("auth_method") == "google",
        tracker_type="calorie"
    )


@calorie_bp.route("/add", methods=["POST"])
@login_required
def add():
    """Add a calorie entry"""
    user_id = session["user_id"]
    food_item = request.form.get("food_item", "").strip()
    quantity = request.form.get("quantity", "1").strip()
    calories = request.form.get("calories", "").strip()
    entry_date = request.form.get("entry_date", date.today().isoformat())

    if not food_item or not calories or not quantity:
        flash("Food item, quantity, and calories are required.", "error")
        return redirect(url_for("calorie.index", date=entry_date))

    try:
        quantity = float(quantity)
        calories = int(calories)
        if quantity <= 0 or calories <= 0:
            raise ValueError
    except ValueError:
        flash("Quantity and calories must be positive numbers.", "error")
        return redirect(url_for("calorie.index", date=entry_date))

    database.add_entry(user_id, entry_date, food_item, calories, quantity=quantity)
    # Invalidate nutrient cache for this date so it gets recalculated
    database.delete_nutrient_data(user_id, entry_date)
    quantity_label = int(quantity) if quantity.is_integer() else quantity
    flash(f"Added {quantity_label} x {food_item} ({calories} kcal)", "success")
    return redirect(url_for("calorie.index", date=entry_date))


@calorie_bp.route("/repeat/<int:entry_id>", methods=["POST"])
@login_required
def repeat_entry(entry_id):
    """Repeat a previously logged calorie entry for a target date."""
    user_id = session["user_id"]
    entry_date = request.form.get("entry_date", date.today().isoformat())
    entry = database.get_entry_by_id(entry_id, user_id)

    if not entry:
        flash("Could not find that food entry.", "error")
        return redirect(url_for("calorie.index", date=entry_date))

    database.add_entry(
        user_id,
        entry_date,
        entry["food_item"],
        entry["calories"],
        quantity=entry.get("quantity", 1),
    )
    database.delete_nutrient_data(user_id, entry_date)
    flash(f"Added {entry['food_item']} again.", "success")
    return redirect(url_for("calorie.index", date=entry_date))


@calorie_bp.route("/delete/<int:entry_id>", methods=["POST"])
@login_required
def delete(entry_id):
    """Delete a calorie entry"""
    user_id = session["user_id"]
    entry_date = request.form.get("entry_date", date.today().isoformat())
    database.delete_entry(entry_id, user_id)
    # Invalidate nutrient cache for this date so it gets recalculated
    database.delete_nutrient_data(user_id, entry_date)
    flash("Entry deleted successfully.", "success")
    return redirect(url_for("calorie.index", date=entry_date))


@calorie_bp.route("/export", methods=["GET"])
@login_required
def export():
    """Export entries to Excel"""
    user_id = session["user_id"]
    all_dates = database.get_all_dates(user_id)

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Calorie Entries"

    # Headers
    from openpyxl.styles import Font, PatternFill
    headers = ["Date", "Food Item", "Quantity", "Calories", "Time"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFA500",
                                end_color="FFA500", fill_type="solid")

    # Data
    row = 2
    for entry_date in all_dates:
        entries = database.get_entries_by_date(user_id, entry_date)
        for entry in entries:
            ws.cell(row=row, column=1, value=entry_date)
            ws.cell(row=row, column=2, value=entry["food_item"])
            ws.cell(row=row, column=3, value=entry.get("quantity", 1))
            ws.cell(row=row, column=4, value=entry["calories"])
            ws.cell(row=row, column=5, value=entry["created_at"][:16])
            row += 1

    # Column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 16

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"calorie_tracker_{date.today()}.xlsx"
    )


def validate_food_input(food_input: str) -> tuple[bool, str | None]:
    """
    Returns (True, None) if valid.
    Returns (False, reason) if invalid.
    """
    FOOD_MIN_LEN = 2
    FOOD_MAX_LEN = 200

    # allowed: letters, digits, spaces, and common food punctuation
    FOOD_ALLOWED_RE = re.compile(r"^[a-zA-Z0-9\s\-\',\.\(\)\/\%°&]+$")

    # must contain at least one letter (rejects "123", "!!!", etc.)
    FOOD_HAS_LETTER_RE = re.compile(r"[a-zA-Z]")

    # obvious non-food / prompt-injection patterns
    FOOD_BLOCKED_RE = re.compile(
        r"(ignore|forget|pretend|you are|act as|system:|<.*?>|SELECT\s|DROP\s|INSERT\s)",
        re.IGNORECASE
    )

    if not food_input or not food_input.strip():
        return False, "Food input cannot be empty."

    s = food_input.strip()

    if len(s) < FOOD_MIN_LEN:
        return False, f"Food input too short (minimum {FOOD_MIN_LEN} characters)."

    if len(s) > FOOD_MAX_LEN:
        return False, f"Food input too long (maximum {FOOD_MAX_LEN} characters)."

    if not FOOD_HAS_LETTER_RE.search(s):
        return False, "Food input must contain at least one letter."

    if not FOOD_ALLOWED_RE.match(s):
        return False, "Food input contains invalid characters."

    if FOOD_BLOCKED_RE.search(s):
        return False, "Food input contains disallowed content."

    return True, None


def normalize_food_name(food_input: str) -> str:
    """Normalize food text and strip leading quantity words when present."""
    normalized_food = re.sub(r'\s+', ' ', food_input.lower().strip())
    normalized_food = re.sub(
        r'^\d+(\.\d+)?\s*(x\s+)?',
        '',
        normalized_food
    )
    return normalized_food.strip()


def normalize_food_key(food_input: str) -> str:
    """Normalize food string for consistent per-unit cache lookups."""
    return normalize_food_name(food_input)


@calorie_bp.route("/suggest-food", methods=["POST"])
@login_required
def suggest_food():
    """AI-powered food suggestion with caching."""
    try:
        food_input = request.json.get("food_input", "").strip()
        quantity_raw = request.json.get("quantity", 1)
        quantity = float(quantity_raw)
        if quantity <= 0:
            raise ValueError("Quantity must be positive.")

        if not food_input:
            return jsonify({"error": "Food input is required"}), 400

        is_valid, reason = validate_food_input(food_input)
        if not is_valid:
            return jsonify({"error": reason, "valid": False}), 422

        food_key = normalize_food_key(food_input)

        # ── 1. Cache lookup (fast path) ────────────────────────────────
        db = database.get_connection()
        cached = db.execute(
            "SELECT food_name, calories FROM food_calorie_cache WHERE food_key = ?",
            (food_key,)
        ).fetchone()

        if cached:
            unit_calories = int(cached["calories"])
            total_calories = int(round(unit_calories * quantity))
            db.execute(
                """UPDATE food_calorie_cache
                   SET hit_count = hit_count + 1,
                       last_accessed = ?
                   WHERE food_key = ?""",
                (datetime.utcnow(), food_key)
            )
            db.commit()
            db.close()
            return jsonify({
                "food": cached["food_name"],
                "calories": total_calories,
                "unit_calories": unit_calories,
                "quantity": quantity,
                "suggestion": f"{cached['food_name']}: {unit_calories} kcal each, {total_calories} kcal total",
                "cached": True
            })

        api_key = Config.OPENAI_API_KEY
        if not api_key:
            return jsonify({"error": "OpenAI API key not configured"}), 500

        client = OpenAI(api_key=api_key)
        quantity_label = int(quantity) if quantity.is_integer() else quantity
        base_food_name = normalize_food_name(food_input)
        response = client.chat.completions.create(
            model=Config.OPENAI_MODEL,
            messages=[
                {
                    "role": "system",
                    "content": "You are a calorie estimation assistant. "
                               "Estimate calories for ONE unit or ONE serving only. "
                               "Do not estimate for multiple pieces together. "
                               "Respond with ONLY: 'food_name: calories_number'."
                },
                {
                    "role": "user",
                    "content": f"Estimate calories for 1 serving/item of {base_food_name}. The user plans to log quantity {quantity_label}, but you must return calories for one item/serving only."
                }
            ],
            temperature=Config.OPENAI_TEMPERATURE,
            max_tokens=Config.OPENAI_MAX_TOKENS
        )

        suggestion = response.choices[0].message.content.strip()

        if ":" not in suggestion:
            db.close()
            return jsonify({
                "error": f"Could not suggest calorie value for {food_input}",
                "raw_response": suggestion
            }), 400

        parts = suggestion.split(":")
        food_name = parts[0].strip()
        try:
            unit_calories = int(parts[-1].strip().split()[0])
        except (ValueError, IndexError):
            db.close()
            return jsonify({
                "error": f"Could not suggest calorie value for {food_input}",
                "raw_response": suggestion
            }), 400

        calories = int(round(unit_calories * quantity))

        db.execute(
            """INSERT INTO food_calorie_cache (food_key, food_name, calories)
               VALUES (?, ?, ?)
               ON CONFLICT(food_key) DO UPDATE SET
                   calories = excluded.calories,
                   food_name = excluded.food_name,
                   hit_count = hit_count + 1,
                   last_accessed = CURRENT_TIMESTAMP""",
            (food_key, food_name, unit_calories)
        )
        db.commit()
        db.close()

        return jsonify({
            "food": food_name,
            "calories": calories,
            "unit_calories": unit_calories,
            "quantity": quantity,
            "suggestion": f"{food_name}: {unit_calories} kcal each, {calories} kcal total for {quantity_label}",
            "cached": False
        })

    except Exception as e:
        print(f"Error in suggest_food: {str(e)}")
        return jsonify({"error": str(e)}), 500


@calorie_bp.route("/nutrient-breakdown", methods=["GET"])
@login_required
def nutrient_breakdown():
    """Get nutrient breakdown - from cache or calculate with AI"""
    try:
        user_id = session["user_id"]
        selected_date = request.args.get("date", date.today().isoformat())
        entries = database.get_entries_by_date(user_id, selected_date)

        if not entries:
            return jsonify({
                "protein": 0,
                "carbs": 0,
                "fats": 0,
                "fiber": 0,
                "details": "No entries for this date",
                "success": True
            })

        # Check if nutrient data already exists in database
        cached_nutrients = database.get_nutrient_data(user_id, selected_date)

        if cached_nutrients:
            # Return from database
            total_cals = sum(e['calories'] for e in entries)
            protein_cals = cached_nutrients['protein_grams'] * 4
            carbs_cals = cached_nutrients['carbs_grams'] * 4
            fats_cals = cached_nutrients['fats_grams'] * 9

            protein_pct = round(
                (protein_cals / total_cals * 100)) if total_cals > 0 else 0
            carbs_pct = round((carbs_cals / total_cals * 100)
                              ) if total_cals > 0 else 0
            fats_pct = round((fats_cals / total_cals * 100)
                             ) if total_cals > 0 else 0
            return jsonify({
                "protein": cached_nutrients['protein_grams'],
                "protein_pct": protein_pct,
                "carbs": cached_nutrients['carbs_grams'],
                "carbs_pct": carbs_pct,
                "fats": cached_nutrients['fats_grams'],
                "fats_pct": fats_pct,
                "fiber": cached_nutrients['fiber_grams'],
                "total_calories": total_cals,
                "analysis": cached_nutrients['analysis'],
                "success": True,
                "from_cache": True
            })

        # No cache found - call OpenAI to analyze nutrients
        # Prepare food list for nutrient analysis
        food_list = "\n".join(
            [
                f"- {e['food_item']} x{int(e['quantity']) if float(e['quantity']).is_integer() else e['quantity']} ({e['calories']} kcal)"
                for e in entries
            ]
        )

        api_key = Config.OPENAI_API_KEY
        if not api_key:
            return jsonify({"error": "OpenAI API key not configured"}), 500

        client = OpenAI(api_key=api_key)

        system_prompt = """You are a nutritionist analyzing daily food intake. Based on the food items and calories,
        estimate the macronutrient breakdown. Respond ONLY with valid JSON (no markdown, no extra text):
        {
            "protein_grams": <number>,
            "carbs_grams": <number>,
            "fats_grams": <number>,
            "fiber_grams": <number>,
            "analysis": "<brief summary>"
        }"""

        response = client.chat.completions.create(
            model=Config.OPENAI_MODEL,
            messages=[
                {
                    "role": "system",
                    "content": system_prompt
                },
                {
                    "role": "user",
                    "content": f"Analyze this day's nutrition:\n{food_list}\nTotal: {sum(e['calories'] for e in entries)} kcal"
                }
            ],
            temperature=Config.OPENAI_TEMPERATURE,
            max_tokens=500
        )

        response_text = response.choices[0].message.content.strip()

        # Clean response if wrapped in markdown code blocks
        if response_text.startswith("```"):
            response_text = response_text.split("```")[1]
            if response_text.startswith("json"):
                response_text = response_text[4:]

        import json
        result = json.loads(response_text)

        # Calculate percentages
        total_cals = sum(e['calories'] for e in entries)
        protein_cals = result.get('protein_grams', 0) * 4
        carbs_cals = result.get('carbs_grams', 0) * 4
        fats_cals = result.get('fats_grams', 0) * 9

        protein_pct = round((protein_cals / total_cals * 100)
                            ) if total_cals > 0 else 0
        carbs_pct = round((carbs_cals / total_cals * 100)
                          ) if total_cals > 0 else 0
        fats_pct = round((fats_cals / total_cals * 100)
                         ) if total_cals > 0 else 0

        # Save nutrients to database for future use
        analysis = result.get('analysis', 'Daily nutrition analyzed')
        database.save_nutrient_data(
            user_id,
            selected_date,
            result.get('protein_grams', 0),
            result.get('carbs_grams', 0),
            result.get('fats_grams', 0),
            result.get('fiber_grams', 0),
            analysis
        )

        return jsonify({
            "protein": result.get('protein_grams', 0),
            "protein_pct": protein_pct,
            "carbs": result.get('carbs_grams', 0),
            "carbs_pct": carbs_pct,
            "fats": result.get('fats_grams', 0),
            "fats_pct": fats_pct,
            "fiber": result.get('fiber_grams', 0),
            "total_calories": total_cals,
            "analysis": analysis,
            "success": True,
            "from_cache": False
        })

    except json.JSONDecodeError as e:
        print(f"JSON parse error: {str(e)}")
        return jsonify({"error": "Failed to parse nutrition data", "success": False}), 500
    except Exception as e:
        print(f"Error in nutrient_breakdown: {str(e)}")
        return jsonify({"error": str(e), "success": False}), 500
